import os
import datetime
from io import BytesIO
from flask import Flask, render_template, request, redirect, url_for, flash, session, make_response
from flask_sqlalchemy import SQLAlchemy
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'super_secret_key_for_flash_messages'

# --- DATABASE CONFIGURATION ---
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'enactus.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# --- MODELS ---
class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    date_day = db.Column(db.String(10), nullable=False) 
    date_month = db.Column(db.String(10), nullable=False) 
    short_desc = db.Column(db.String(200), nullable=False)
    full_desc = db.Column(db.Text, nullable=False)
    image_url = db.Column(db.String(500), nullable=False)
    is_open = db.Column(db.Boolean, default=True)
    # Event type: 'solo' or 'team'
    event_type = db.Column(db.String(20), default='solo')  # 'solo' or 'team'
    # New fields for event duration
    start_date = db.Column(db.String(20), nullable=True)  # Format: YYYY-MM-DD
    end_date = db.Column(db.String(20), nullable=True)    # Format: YYYY-MM-DD
    event_time = db.Column(db.String(20), nullable=True)  # Format: HH:MM
    venue = db.Column(db.String(200), nullable=True)
    max_registrations = db.Column(db.Integer, nullable=True)
    # Team event fields
    min_team_size = db.Column(db.Integer, nullable=True)  # Minimum team members
    max_team_size = db.Column(db.Integer, nullable=True)  # Maximum team members
    event_link = db.Column(db.String(500), nullable=True)  # External link/URL for event

class Registration(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    # Common fields
    registration_type = db.Column(db.String(20), default='solo')  # 'solo' or 'team'
    # Solo event fields
    name = db.Column(db.String(100), nullable=True)
    email = db.Column(db.String(100), nullable=True)
    student_id = db.Column(db.String(50), nullable=True)
    contact_no = db.Column(db.String(20), nullable=True)
    branch = db.Column(db.String(100), nullable=True)
    college_name = db.Column(db.String(200), nullable=True)
    # Team event fields
    team_name = db.Column(db.String(100), nullable=True)
    team_size = db.Column(db.Integer, nullable=True)
    leader_name = db.Column(db.String(100), nullable=True)
    leader_email = db.Column(db.String(100), nullable=True)
    leader_contact = db.Column(db.String(20), nullable=True)
    # Relationship
    event = db.relationship('Event', backref='registrations')

# --- PUBLIC ROUTES ---
@app.route('/')
def home(): return render_template('home.html', title="Home")

@app.route('/projects')
def projects(): return render_template('projects.html', title="Our Work")

@app.route('/project/navodaya')
def project_navodaya(): return render_template('project_navodaya.html', title="Project Navodaya")

@app.route('/project/astitva')
def project_astitva(): return render_template('project_astitva.html', title="Project Astitva")

@app.route('/project/vriksh')
def project_vriksh(): return render_template('project_vriksh.html', title="Project Vriksh")

@app.route('/team')
def team(): return render_template('team.html', title="Our Team")

@app.route('/about')
def about(): return render_template('about.html', title="About Us")

@app.route('/store')
def store(): return render_template('store.html', title="Store")

@app.route('/events')
def events():
    all_events = Event.query.all()
    return render_template('events.html', events=all_events, title="Events")

@app.route('/register_event', methods=['POST'])
def register_event():
    event_id = request.form.get('event_id')
    event_type = request.form.get('event_type', 'solo')
    
    # Check event exists and is open
    event = Event.query.get(event_id)
    if not event:
        flash("Event not found.", "error")
        return redirect(url_for('events'))
    
    if not event.is_open:
        flash("Registration for this event is closed.", "error")
        return redirect(url_for('events'))
    
    # Check max registrations limit
    if event.max_registrations:
        current_reg_count = Registration.query.filter_by(event_id=event_id).count()
        if current_reg_count >= event.max_registrations:
            flash("Sorry, this event is full. Registration is closed.", "error")
            return redirect(url_for('events'))

    try:
        if event_type == 'team':
            # Team event registration
            team_name = request.form.get('team_name')
            team_size = request.form.get('team_size')
            leader_name = request.form.get('name')
            leader_email = request.form.get('email')
            leader_contact = request.form.get('phone')
            college = request.form.get('college')
            
            if not team_name or not leader_name or not leader_email:
                flash("Please fill in all required fields.", "error")
                return redirect(url_for('events'))
            
            new_reg = Registration(
                event_id=event_id,
                registration_type='team',
                team_name=team_name,
                team_size=int(team_size) if team_size else None,
                leader_name=leader_name,
                leader_email=leader_email,
                leader_contact=leader_contact,
                college_name=college,
                name=leader_name,
                email=leader_email
            )
            reg_name = team_name
        else:
            # Solo event registration
            name = request.form.get('name')
            email = request.form.get('email')
            student_id = request.form.get('student_id')
            phone = request.form.get('phone')
            branch = request.form.get('branch')
            college = request.form.get('college')
            
            if not name or not email:
                flash("Please fill in all required fields.", "error")
                return redirect(url_for('events'))
            
            new_reg = Registration(
                event_id=event_id,
                registration_type='solo',
                name=name,
                email=email,
                student_id=student_id,
                contact_no=phone,
                branch=branch,
                college_name=college
            )
            reg_name = name
        
        db.session.add(new_reg)
        db.session.commit()
        
        # Check if max reached after this registration and auto-close if so
        if event.max_registrations:
            updated_count = Registration.query.filter_by(event_id=event_id).count()
            if updated_count >= event.max_registrations:
                # Event is full but keeps is_open=True so it shows in Upcoming (as "Full")
                flash(f"Success! {reg_name}, you are registered. The event has now reached maximum capacity.", "success")
            else:
                flash(f"Success! {reg_name}, you are registered. ({updated_count}/{event.max_registrations} spots filled)", "success")
        else:
            flash(f"Success! {reg_name}, you are registered.", "success")
    except Exception as e:
        db.session.rollback()
        print(f"Registration error: {e}")
        flash("Something went wrong.", "error")

    return redirect(url_for('events'))

# --- ADMIN ROUTES (SECURE) ---

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password')
        # Simple Password Check
        if password == 'enactus_adgips':
            session['is_admin'] = True
            flash("Welcome back, Admin.", "success")
            return redirect(url_for('admin'))
        else:
            flash("Invalid Password.", "error")
            return redirect(url_for('admin_login'))
            
    return render_template('admin_login.html', title="Admin Login")

@app.route('/admin')
def admin():
    # Security Check
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))

    regs = Registration.query.all()
    events = Event.query.all()
    return render_template('admin.html', regs=regs, events=events, title="Admin Dashboard")

@app.route('/add_event', methods=['POST'])
def add_event():
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))

    title = request.form.get('title')
    start_date = request.form.get('start_date')
    end_date = request.form.get('end_date')
    short_desc = request.form.get('short_desc')
    full_desc = request.form.get('full_desc')
    image_url = request.form.get('image_url')
    # Event type
    event_type = request.form.get('event_type', 'solo')
    # Team size constraints
    min_team_size = request.form.get('min_team_size')
    max_team_size = request.form.get('max_team_size')
    # Other fields
    event_time = request.form.get('time')
    venue = request.form.get('venue')
    max_reg = request.form.get('max_registrations')
    event_link = request.form.get('event_link')

    try:
        start_date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
        day = start_date_obj.strftime("%d")
        month = start_date_obj.strftime("%b").upper()

        new_event = Event(
            title=title,
            date_day=day,
            date_month=month,
            short_desc=short_desc,
            full_desc=full_desc,
            image_url=image_url,
            is_open=True,
            event_type=event_type,
            start_date=start_date,
            end_date=end_date if end_date else start_date,
            event_time=event_time if event_time else None,
            venue=venue if venue else None,
            max_registrations=int(max_reg) if max_reg else None,
            min_team_size=int(min_team_size) if min_team_size and event_type == 'team' else None,
            max_team_size=int(max_team_size) if max_team_size and event_type == 'team' else None,
            event_link=event_link if event_link else None
        )
        db.session.add(new_event)
        db.session.commit()
        flash("New Event Published Successfully!", "success")
    except Exception as e:
        print(e)
        flash("Error creating event. Check date format.", "error")

    return redirect(url_for('admin'))

@app.route('/edit_event/<int:event_id>', methods=['GET', 'POST'])
def edit_event(event_id):
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    event = Event.query.get_or_404(event_id)
    
    if request.method == 'POST':
        event.title = request.form.get('title')
        event.short_desc = request.form.get('short_desc')
        event.full_desc = request.form.get('full_desc')
        event.image_url = request.form.get('image_url')
        event.venue = request.form.get('venue') if request.form.get('venue') else None
        event.event_time = request.form.get('time') if request.form.get('time') else None
        
        # Handle event type
        event.event_type = request.form.get('event_type', 'solo')
        
        # Handle team size for team events
        min_team = request.form.get('min_team_size')
        max_team = request.form.get('max_team_size')
        event.min_team_size = int(min_team) if min_team and event.event_type == 'team' else None
        event.max_team_size = int(max_team) if max_team and event.event_type == 'team' else None
        
        # Handle dates
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        if start_date:
            event.start_date = start_date
            date_obj = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            event.date_day = date_obj.strftime("%d")
            event.date_month = date_obj.strftime("%b").upper()
        if end_date:
            event.end_date = end_date
        else:
            event.end_date = start_date
            
        # Handle max registrations
        max_reg = request.form.get('max_registrations')
        event.max_registrations = int(max_reg) if max_reg else None
        
        # Handle event link
        event_link = request.form.get('event_link')
        event.event_link = event_link if event_link else None
        
        db.session.commit()
        flash(f"Event '{event.title}' updated successfully!", "success")
        return redirect(url_for('admin'))
    
    return render_template('edit_event.html', event=event, title="Edit Event")

@app.route('/toggle_event/<int:event_id>', methods=['POST'])
def toggle_event(event_id):
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    event = Event.query.get_or_404(event_id)
    event.is_open = not event.is_open
    db.session.commit()
    
    status = "opened" if event.is_open else "closed"
    flash(f"Registration for '{event.title}' has been {status}.", "success")
    return redirect(url_for('admin'))

@app.route('/delete_event/<int:event_id>', methods=['POST'])
def delete_event(event_id):
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    event = Event.query.get_or_404(event_id)
    event_title = event.title
    
    # Delete all registrations for this event first
    Registration.query.filter_by(event_id=event_id).delete()
    
    db.session.delete(event)
    db.session.commit()
    
    flash(f"Event '{event_title}' and all its registrations have been deleted.", "success")
    return redirect(url_for('admin'))

@app.route('/delete_registration/<int:reg_id>', methods=['POST'])
def delete_registration(reg_id):
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    registration = Registration.query.get_or_404(reg_id)
    reg_name = registration.name
    
    db.session.delete(registration)
    db.session.commit()
    
    flash(f"Registration for '{reg_name}' has been deleted.", "success")
    return redirect(url_for('admin'))

@app.route('/logout')
def logout():
    session.pop('is_admin', None)
    flash("Logged out successfully.", "success")
    return redirect(url_for('home'))

# --- PDF EXPORT ROUTES ---
@app.route('/export_all_pdf')
def export_all_pdf():
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, spaceAfter=20, textColor=colors.HexColor('#d4a553'))
    elements.append(Paragraph("Enactus - All Registrations", title_style))
    elements.append(Spacer(1, 0.25*inch))
    
    # Get all events with registrations
    events = Event.query.all()
    
    for event in events:
        regs = Registration.query.filter_by(event_id=event.id).all()
        if not regs:
            continue
            
        # Event header
        event_style = ParagraphStyle('EventTitle', parent=styles['Heading2'], fontSize=14, spaceAfter=10, spaceBefore=20)
        elements.append(Paragraph(f"{event.title} ({event.date_day} {event.date_month})", event_style))
        
        # Table data
        if event.event_type == 'team':
            headers = ['#', 'Team Name', 'Leader Name', 'Leader Email', 'Leader Contact', 'Team Size']
            data = [headers]
            for i, reg in enumerate(regs, 1):
                data.append([str(i), reg.team_name or '-', reg.leader_name or '-', reg.leader_email or '-', reg.leader_contact or '-', str(reg.team_size or '-')])
        else:
            headers = ['#', 'Name', 'Email', 'Student ID', 'Contact', 'Branch', 'College']
            data = [headers]
            for i, reg in enumerate(regs, 1):
                data.append([str(i), reg.name or '-', reg.email or '-', reg.student_id or '-', reg.contact_no or '-', reg.branch or '-', reg.college_name or '-'])
        
        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1f26')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#d4a553')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f8f8')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=all_registrations.pdf'
    return response

@app.route('/export_event_pdf/<int:event_id>')
def export_event_pdf(event_id):
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    event = Event.query.get_or_404(event_id)
    regs = Registration.query.filter_by(event_id=event_id).all()
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, spaceAfter=10, textColor=colors.HexColor('#d4a553'))
    elements.append(Paragraph(f"Enactus - {event.title}", title_style))
    
    # Event info
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=11, spaceAfter=20)
    elements.append(Paragraph(f"Date: {event.date_day} {event.date_month} | Venue: {event.venue or 'TBA'} | Total Registrations: {len(regs)}", info_style))
    elements.append(Spacer(1, 0.25*inch))
    
    if regs:
        # Table data
        if event.event_type == 'team':
            headers = ['#', 'Team Name', 'Leader Name', 'Leader Email', 'Leader Contact', 'Team Size']
            data = [headers]
            for i, reg in enumerate(regs, 1):
                data.append([str(i), reg.team_name or '-', reg.leader_name or '-', reg.leader_email or '-', reg.leader_contact or '-', str(reg.team_size or '-')])
        else:
            headers = ['#', 'Name', 'Email', 'Student ID', 'Contact', 'Branch', 'College']
            data = [headers]
            for i, reg in enumerate(regs, 1):
                data.append([str(i), reg.name or '-', reg.email or '-', reg.student_id or '-', reg.contact_no or '-', reg.branch or '-', reg.college_name or '-'])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1f26')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#d4a553')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f8f8')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cccccc')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No registrations yet.", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={event.title.replace(" ", "_")}_registrations.pdf'
    return response

# --- EXCEL EXPORT ROUTES ---
@app.route('/export_all_excel')
def export_all_excel():
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    if not OPENPYXL_AVAILABLE:
        flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'error')
        return redirect(url_for('admin'))
    
    wb = Workbook()
    events = Event.query.all()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3c34", end_color="1a3c34", fill_type="solid")
    
    for event in events:
        regs = Registration.query.filter_by(event_id=event.id).all()
        if not regs:
            continue
        
        # Create sheet for each event (limit title to 30 chars)
        ws = wb.create_sheet(title=event.title[:30])
        
        if event.event_type == 'team':
            headers = ['#', 'Team Name', 'Leader Name', 'Leader Email', 'Leader Contact', 'Team Size', 'College']
            ws.append(headers)
            for i, reg in enumerate(regs, 1):
                ws.append([i, reg.team_name or '', reg.leader_name or '', reg.leader_email or '', reg.leader_contact or '', reg.team_size or '', reg.college_name or ''])
        else:
            headers = ['#', 'Name', 'Email', 'Student ID', 'Contact', 'Branch', 'College']
            ws.append(headers)
            for i, reg in enumerate(regs, 1):
                ws.append([i, reg.name or '', reg.email or '', reg.student_id or '', reg.contact_no or '', reg.branch or '', reg.college_name or ''])
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)
    
    if not wb.worksheets:
        ws = wb.create_sheet(title="No Data")
        ws.append(["No registrations found"])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=all_registrations.xlsx'
    return response


@app.route('/export_event_excel/<int:event_id>')
def export_event_excel(event_id):
    if not session.get('is_admin'):
        return redirect(url_for('admin_login'))
    
    if not OPENPYXL_AVAILABLE:
        flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'error')
        return redirect(url_for('admin'))
    
    event = Event.query.get_or_404(event_id)
    regs = Registration.query.filter_by(event_id=event_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = event.title[:30]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a3c34", end_color="1a3c34", fill_type="solid")
    
    if event.event_type == 'team':
        headers = ['#', 'Team Name', 'Leader Name', 'Leader Email', 'Leader Contact', 'Team Size', 'College']
        ws.append(headers)
        for i, reg in enumerate(regs, 1):
            ws.append([i, reg.team_name or '', reg.leader_name or '', reg.leader_email or '', reg.leader_contact or '', reg.team_size or '', reg.college_name or ''])
    else:
        headers = ['#', 'Name', 'Email', 'Student ID', 'Contact', 'Branch', 'College']
        ws.append(headers)
        for i, reg in enumerate(regs, 1):
            ws.append([i, reg.name or '', reg.email or '', reg.student_id or '', reg.contact_no or '', reg.branch or '', reg.college_name or ''])
    
    # Style headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={event.title.replace(" ", "_")}_registrations.xlsx'
    return response

def seed_database():
    with app.app_context():
        db.create_all()
        if Event.query.count() == 0:
            e1 = Event(
                title="Social Entrepreneurship Summit",
                date_day="12", date_month="NOV",
                short_desc="Leading the future of social impact.",
                full_desc="Join industry giants for a masterclass.",
                image_url="https://images.unsplash.com/photo-1544531586-fde5298cdd40?w=800",
                is_open=True
            )
            db.session.add(e1)
            db.session.commit()

if __name__ == '__main__':
    seed_database()
    app.run(host='0.0.0.0',port=5003, threaded=True)
